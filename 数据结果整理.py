import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 加载Excel数据
data = pd.read_excel(r'C:\Users\10133\Desktop\evrp\改\meta_heuristics_results with 1s and 30threshold.xlsx')

# 按实例文件名和算法进行分组，计算每组目标函数（D列）的均值、标准差和最小值
result = data.groupby(['File', 'MetaHeuristic'])['Cost'].agg(['mean', 'std', 'min']).reset_index()

# 保存结果为Excel
result_file = '1s and 30threshold.xlsx'
result.to_excel(result_file, index=False)

# 重新加载保存的Excel文件
wb = load_workbook(result_file)
ws = wb.active

# 查找每个实例的最小值并加粗
instances = result['File'].unique()

for instance in instances:
    # 获取同一实例的行索引
    instance_rows = result[result['File'] == instance]

    # 获取最小的目标函数值和最小值
    min_mean = instance_rows['mean'].min()
    min_min = instance_rows['min'].min()

    # 忽略FCFS，获取标准差最小值
    std_filtered = instance_rows[instance_rows['MetaHeuristic'] != 'FCFS']
    min_std = std_filtered['std'].min() if not std_filtered.empty else None

    # 循环检查对应列的值，并在Excel中加粗最小值
    for row in instance_rows.index:
        excel_row = row + 2  # 对应Excel行，+2是因为pandas的索引从0开始，Excel从第2行开始有数据

        # 加粗最小的 mean
        if result.at[row, 'mean'] == min_mean:
            ws.cell(row=excel_row, column=3).font = Font(color="FF0000")

        # 加粗最小的 std（忽略FCFS）
        if result.at[row, 'std'] == min_std and result.at[row, 'MetaHeuristic'] != 'FCFS':
            ws.cell(row=excel_row, column=4).font = Font(color="FF0000")

        # 加粗最小的 min
        if result.at[row, 'min'] == min_min:
            ws.cell(row=excel_row, column=5).font = Font(color="FF0000")

# 保存最终文件
wb.save(result_file)

print("结果已保存并加粗最小值。")